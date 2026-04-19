from __future__ import annotations

import logging
import os
from pathlib import Path
from typing import Any

import polars as pl

logger = logging.getLogger(__name__)

# Polars dtypes considered numeric
_NUMERIC_DTYPES = (
    pl.Float32, pl.Float64,
    pl.Int8, pl.Int16, pl.Int32, pl.Int64,
    pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64,
)
_DATE_DTYPES = (pl.Date, pl.Datetime, pl.Duration)

# Heuristic thresholds
_LOW_CARD_MAX_RATIO = 0.05   # unique / total ≤ 5% → attribute
_MEASURE_MIN_UNIQUE = 10      # numeric with more than 10 unique values → measure
_SAMPLE_SIZE = 5              # values to include in sample_values


def _infer_role(col_name: str, series: pl.Series, total_rows: int) -> str:
    """Heuristically infer column_role."""
    dtype = series.dtype

    if isinstance(dtype, _DATE_DTYPES):
        return "time"

    if isinstance(dtype, _NUMERIC_DTYPES):
        unique_count = series.n_unique()
        # Columns named like *id*, *key*, *code*, *number* → key even if numeric
        lower = col_name.lower()
        if any(tok in lower for tok in ("_id", "id_", " id", "key", "_nr", "number", "code")):
            return "key"
        if unique_count >= _MEASURE_MIN_UNIQUE:
            return "measure"
        return "attribute"

    # String / categorical columns
    unique_count = series.n_unique()
    lower = col_name.lower()
    if any(tok in lower for tok in ("_id", "id_", " id", "key", "_nr", "number", "code")):
        return "key"
    if total_rows > 0 and (unique_count / total_rows) <= _LOW_CARD_MAX_RATIO:
        return "attribute"
    return "attribute"


def _polars_dtype_to_str(dtype: pl.PolarsDataType) -> str:
    """Map a Polars dtype to a canonical string for metadata."""
    if isinstance(dtype, pl.Datetime):
        return "date"
    if isinstance(dtype, pl.Date):
        return "date"
    if dtype in (pl.Float32, pl.Float64):
        return "numeric"
    if dtype in (pl.Int8, pl.Int16, pl.Int32, pl.Int64,
                 pl.UInt8, pl.UInt16, pl.UInt32, pl.UInt64):
        return "integer"
    if dtype == pl.Boolean:
        return "boolean"
    # String / Categorical / Utf8View all map to text
    return "text"


def _detect_currency(col_name: str, series: pl.Series) -> bool:
    """Heuristic: name contains currency keywords and data is numeric."""
    lower = col_name.lower()
    currency_tokens = (
        "amount", "betrag", "value", "wert", "saldo",
        "revenue", "cost", "price", "preis", "umsatz",
        "total", "sum", "summe",
    )
    return any(t in lower for t in currency_tokens) and isinstance(series.dtype, _NUMERIC_DTYPES)


def _normalize_schema(df: pl.DataFrame) -> pl.DataFrame:
    """Replace Utf8View and other non-standard dtypes with canonical equivalents."""
    casts: list[pl.Expr] = []
    for col_name, dtype in zip(df.columns, df.dtypes):
        dtype_str = str(dtype)
        # Utf8View → String
        if dtype_str in ("Utf8View", "String(View)"):
            casts.append(pl.col(col_name).cast(pl.String))
        # Large String → String
        elif dtype_str in ("LargeUtf8", "LargeString"):
            casts.append(pl.col(col_name).cast(pl.String))
        else:
            casts.append(pl.col(col_name))
    if casts:
        df = df.select(casts)
    return df


def _sample_values(series: pl.Series, n: int = _SAMPLE_SIZE) -> list[Any]:
    """Return up to n non-null sample values as JSON-safe Python scalars."""
    non_null = series.drop_nulls()
    taken = non_null.head(n).to_list()
    # Convert date/datetime/timedelta to strings for JSONB storage
    return [v.isoformat() if hasattr(v, 'isoformat') else str(v) if not isinstance(v, (int, float, str, bool)) else v for v in taken]


_EXCEL_SUFFIXES = (".xlsx", ".xls", ".xlsm", ".xlsb", ".ods")


def list_excel_sheets(file_path: str) -> list[str]:
    """Return the ordered list of sheet names in an Excel workbook.

    Returns an empty list for non-Excel files, or when no engine can read the
    workbook. Callers can then fall back to single-dataset behaviour.
    """
    path = Path(file_path)
    if path.suffix.lower() not in _EXCEL_SUFFIXES:
        return []

    def _clean(raw: list) -> list[str]:
        out: list[str] = []
        for item in raw:
            if item is None:
                continue
            s = str(item).strip()
            if s:
                out.append(s)
        return out

    # fastexcel (calamine binding) enumerates sheets without loading data.
    try:
        import fastexcel

        reader = fastexcel.read_excel(str(path))
        names = _clean(list(reader.sheet_names))
        if names:
            logger.info("fastexcel detected %d sheet(s) in %s: %s", len(names), file_path, names)
            return names
    except Exception as exc:  # pragma: no cover - runtime-dep dependent
        logger.warning("fastexcel sheet enumeration failed for %s: %s", file_path, exc)

    # Fallback: openpyxl read-only mode for .xlsx.
    try:
        from openpyxl import load_workbook

        wb = load_workbook(str(path), read_only=True, data_only=True)
        try:
            names = _clean(list(wb.sheetnames))
            if names:
                logger.info("openpyxl detected %d sheet(s) in %s: %s", len(names), file_path, names)
                return names
        finally:
            wb.close()
    except Exception as exc:  # pragma: no cover
        logger.warning("openpyxl sheet enumeration failed for %s: %s", file_path, exc)

    # Last resort: ask Polars to load every sheet (expensive but reliable).
    try:
        data = pl.read_excel(str(path), engine="calamine", sheet_id=0, infer_schema_length=0)
        if isinstance(data, dict):
            names = _clean(list(data.keys()))
            if names:
                logger.info("polars detected %d sheet(s) in %s: %s", len(names), file_path, names)
                return names
    except Exception as exc:
        logger.warning("polars sheet enumeration failed for %s: %s", file_path, exc)

    return []


def parse_file(
    file_path: str,
    sheet_name: str | None = None,
) -> tuple[pl.DataFrame, list[dict]]:
    """Parse xlsx/csv into DataFrame + column metadata list.

    Column metadata items::

        {
            source_name: str,
            display_name: str,
            data_type: str,        # text|numeric|integer|date|boolean|currency
            column_role: str,      # key|measure|time|attribute
            unique_count: int,
            sample_values: list,
        }

    When ``sheet_name`` is given (Excel only) that sheet is read; otherwise the
    first sheet is used. CSV/TSV inputs ignore ``sheet_name``.

    Uses calamine engine for Excel. Detects types via Polars dtype inspection.
    Normalises Utf8View → String before processing.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")

    suffix = path.suffix.lower()
    logger.info(
        "Parsing file %s (type=%s sheet=%s)", file_path, suffix, sheet_name or "<default>"
    )

    if suffix in _EXCEL_SUFFIXES:
        kwargs: dict[str, Any] = {"infer_schema_length": 1000}
        if sheet_name:
            kwargs["sheet_name"] = sheet_name
        try:
            df = pl.read_excel(file_path, engine="calamine", **kwargs)
        except Exception:
            logger.warning("calamine failed, falling back to openpyxl for %s", file_path)
            df = pl.read_excel(file_path, engine="openpyxl", **kwargs)
    elif suffix == ".csv":
        df = pl.read_csv(
            file_path,
            infer_schema_length=1000,
            ignore_errors=True,
            try_parse_dates=True,
        )
    elif suffix == ".tsv":
        df = pl.read_csv(
            file_path,
            separator="\t",
            infer_schema_length=1000,
            ignore_errors=True,
            try_parse_dates=True,
        )
    else:
        raise ValueError(f"Unsupported file type: {suffix}")

    # Normalise schema
    df = _normalize_schema(df)

    total_rows = len(df)
    logger.info("Parsed %d rows × %d columns", total_rows, len(df.columns))

    column_metadata: list[dict] = []
    for col_name in df.columns:
        series = df[col_name]
        dtype = series.dtype
        unique_count = series.n_unique()

        # Determine data_type string
        data_type = _polars_dtype_to_str(dtype)
        if data_type in ("numeric", "integer") and _detect_currency(col_name, series):
            data_type = "currency"

        column_role = _infer_role(col_name, series, total_rows)
        # Override: if data_type is currency, force measure
        if data_type == "currency":
            column_role = "measure"

        display_name = col_name.replace("_", " ").strip().title()

        meta: dict = {
            "source_name": col_name,
            "display_name": display_name,
            "data_type": data_type,
            "column_role": column_role,
            "unique_count": unique_count,
            "sample_values": _sample_values(series),
        }
        column_metadata.append(meta)

    return df, column_metadata
